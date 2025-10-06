
from DrissionPage import ChromiumPage
import pandas as pd
from openpyxl import Workbook  
import time
import os

# 询问目标用户数量
print("=== TikTok 粉丝爬虫 ===")
target_count = int(input("请输入您想要爬取的用户数量: "))
print(f"目标用户数量: {target_count}")

# 询问保存路径
print("\n请输入Excel文件保存路径:")
save_directory = input("保存目录 (直接按回车使用当前目录): ").strip()
if not save_directory:
    save_directory = os.getcwd()
    
# 确保目录存在
if not os.path.exists(save_directory):
    try:
        os.makedirs(save_directory)
        print(f"已创建目录: {save_directory}")
    except Exception as e:
        print(f"创建目录失败: {e}")
        save_directory = os.getcwd()
        print(f"使用当前目录: {save_directory}")

# 设置Excel文件完整路径
excel_file_path = os.path.join(save_directory, 'tiktok.xlsx')
print(f"Excel文件将保存到: {excel_file_path}")

dp = ChromiumPage()
print("正在访问TikTok页面...")
dp.get('https://www.tiktok.com/@joyschinesehub')

# 等待页面加载完成
time.sleep(3)

# 方法1: 通过data-e2e属性定位（推荐）
try:
    followers_element = dp.ele('css:span[data-e2e="followers"]')
    if followers_element:
        print("找到粉丝元素，准备点击...")
        followers_element.click()
        print("成功点击粉丝按钮")
    else:
        print("未找到粉丝元素")
except Exception as e:
    print(f"方法1失败: {e}")
    
    # 方法2: 通过文本内容定位
    try:
        followers_element = dp.ele('text:粉丝')
        if followers_element:
            print("通过文本找到粉丝元素，准备点击...")
            followers_element.click()
            print("成功点击粉丝按钮")
        else:
            print("通过文本未找到粉丝元素")
    except Exception as e:
        print(f"方法2失败: {e}")
        
        # 方法3: 通过CSS类名定位（不太稳定，因为类名可能变化）
        try:
            followers_element = dp.ele('css:.css-17j1vt2-5e6d46e3--SpanUnit')
            if followers_element and '粉丝' in followers_element.text:
                print("通过CSS类名找到粉丝元素，准备点击...")
                followers_element.click()
                print("成功点击粉丝按钮")
            else:
                print("通过CSS类名未找到正确的粉丝元素")
        except Exception as e:
            print(f"方法3失败: {e}")

# 点击后等待粉丝列表加载
time.sleep(3)
print("粉丝列表已打开，开始获取用户信息...")

# 初始化用户列表和计数器
all_users = []
collected_usernames = set()  # 用于去重
scroll_count = 0
max_scrolls = 100  # 最大滚动次数，防止无限循环

print(f"开始收集用户信息，目标: {target_count} 个用户")
print("-" * 50)

# 获取粉丝列表容器元素（用于滚动）
try:
    user_list_container = dp.ele('css:.css-1sko41r-5e6d46e3--DivUserListContainer')
    if not user_list_container:
        print("⚠️ 未找到粉丝列表容器，使用页面滚动")
        user_list_container = None
except Exception as e:
    print(f"获取粉丝列表容器失败: {e}，使用页面滚动")
    user_list_container = None

while len(all_users) < target_count and scroll_count < max_scrolls:
    try:
        # 获取当前页面的用户信息
        nicknames = dp.eles('css:span.css-spk7wm-5e6d46e3--SpanNickname')
        usernames = dp.eles('css:p.css-1fl8ity-5e6d46e3--PUniqueId')
        
        current_batch_count = 0
        min_count = min(len(nicknames), len(usernames))
        
        print(f"当前页面找到 {len(nicknames)} 个昵称，{len(usernames)} 个用户名")
        
        for i in range(min_count):
            try:
                nickname = nicknames[i].text.strip()
                username = usernames[i].text.strip()
                
                # 检查是否已经收集过这个用户（去重）
                if username not in collected_usernames:
                    user_info = {
                        '序号': len(all_users) + 1,
                        '昵称': nickname,
                        '用户名': username
                    }
                    all_users.append(user_info)
                    collected_usernames.add(username)
                    current_batch_count += 1
                    
                    print(f"用户 {len(all_users)}: 昵称='{nickname}', 用户名='{username}'")
                    
                    # 如果已达到目标数量，退出循环
                    if len(all_users) >= target_count:
                        print(f"\n✅ 已收集到目标数量 {target_count} 个用户!")
                        break
                        
            except Exception as e:
                print(f"处理第 {i+1} 个用户时出错: {e}")
                continue
        
        print(f"第 {scroll_count + 1} 次滚动，本次新增 {current_batch_count} 个用户，总计 {len(all_users)} 个用户")
        
        # 如果已达到目标数量，退出循环
        if len(all_users) >= target_count:
            break
            
        # 如果连续几次滚动都没有获取到新用户，可能已经到底了
        if current_batch_count == 0:
            print("本次滚动未获取到新用户...")
            if scroll_count > 5:  # 连续5次没有新用户就退出
                print("⚠️ 连续多次滚动未获取到新用户，可能已到列表底部")
                break
        
        # 滚动加载更多用户
        scroll_count += 1
        print(f"🔄 第 {scroll_count} 次滚动，在容器内加载更多用户...")
        
        # 在粉丝列表容器内滚动
        if user_list_container:
            # 方法1: 在容器内滚动到底部
            try:
                user_list_container.scroll.to_bottom()
                print("✅ 在粉丝列表容器内滚动到底部")
            except Exception as e:
                print(f"容器滚动失败，尝试其他方法: {e}")
                # 方法2: 使用JavaScript滚动容器
                try:
                    dp.run_js('document.querySelector(".css-1sko41r-5e6d46e3--DivUserListContainer").scrollTop = document.querySelector(".css-1sko41r-5e6d46e3--DivUserListContainer").scrollHeight')
                    print("✅ 使用JavaScript在容器内滚动")
                except Exception as js_e:
                    print(f"JavaScript滚动也失败: {js_e}")
                    # 方法3: 页面滚动作为备用
                    dp.scroll.to_bottom()
                    print("使用页面滚动作为备用方法")
        else:
            # 如果没有找到容器，使用页面滚动
            dp.scroll.to_bottom()
            print("使用页面滚动")
        
        time.sleep(2)  # 等待新内容加载
        
        # 每隔一段时间显示进度
        if scroll_count % 5 == 0:
            print(f"📊 进度报告: 已滚动 {scroll_count} 次，收集了 {len(all_users)} 个用户 (目标: {target_count})")
            
    except Exception as e:
        print(f"滚动第 {scroll_count + 1} 次时发生错误: {e}")
        scroll_count += 1
        time.sleep(2)
        continue

# 最终结果统计
print("\n" + "=" * 60)
print(f"🎉 爬取完成!")
print(f"目标用户数量: {target_count}")
print(f"实际获取数量: {len(all_users)}")
print(f"总滚动次数: {scroll_count}")
print(f"去重用户名数量: {len(collected_usernames)}")

if len(all_users) < target_count:
    print(f"⚠️ 注意: 实际获取数量少于目标数量，可能是因为:")
    print("   1. 该账号粉丝总数不足")
    print("   2. 网络加载问题")
    print("   3. TikTok限制了数据显示")
    print("   4. 已到达粉丝列表底部")

print("\n📋 完整用户列表:")
print("-" * 60)
for user in all_users:
    print(f"{user['序号']:3d}. 昵称: {user['昵称']:<30} 用户名: {user['用户名']}")

# 保存到Excel文件
print(f"\n💾 正在保存数据到Excel文件...")
try:
    # 创建新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "TikTok粉丝数据"
    
    # 设置表头
    headers = ['序号', '昵称', '用户名', '用户链接', '爬取时间']
    ws.append(headers)
    
    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 添加数据
    import datetime
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    for user in all_users:
        row_data = [
            user['序号'],
            user['昵称'],
            user['用户名'],
            f"https://www.tiktok.com/@{user['用户名']}",
            current_time
        ]
        ws.append(row_data)
    
    # 调整列宽
    column_widths = [8, 35, 25, 50, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # 设置数据行样式
    for row_num in range(2, len(all_users) + 2):
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # 交替行颜色
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 保存文件
    wb.save(excel_file_path)
    print(f"✅ 数据已成功保存到: {excel_file_path}")
    print(f"📊 共保存 {len(all_users)} 条用户记录")
    
    # 尝试打开Excel文件所在目录
    try:
        import subprocess
        subprocess.run(f'explorer /select,"{excel_file_path}"', shell=True)
        print("📂 已自动打开文件所在目录")
    except Exception as e:
        print(f"无法自动打开目录: {e}")
        
except Exception as e:
    print(f"❌ 保存Excel文件失败: {e}")
    print("尝试保存为CSV格式...")
    
    # 备用方案：保存为CSV
    try:
        csv_file_path = os.path.join(save_directory, 'tiktok_backup.csv')
        df = pd.DataFrame(all_users)
        df['用户链接'] = df['用户名'].apply(lambda x: f"https://www.tiktok.com/@{x}")
        df['爬取时间'] = current_time
        df.to_csv(csv_file_path, index=False, encoding='utf-8-sig')
        print(f"✅ 数据已保存为CSV文件: {csv_file_path}")
    except Exception as csv_e:
        print(f"❌ CSV保存也失败: {csv_e}")

print("操作完成")

